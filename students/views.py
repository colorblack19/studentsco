
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponseRedirect, FileResponse, JsonResponse
from django.urls import reverse
from django.db.models import Sum
from django.conf import settings
import os
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch

from .models import Student, Payment, FeeStructure
from .pdf_utils import generate_receipt_pdf, generate_student_statement_pdf  # ← IMPORTANT
from django.contrib.auth.decorators import login_required
from django.contrib import messages
import json
from django.db.models import Sum, Q
from datetime import datetime
from django.contrib.auth.decorators import user_passes_test

from .forms import StudentForm
from .models import FeeStructure 
from .forms import FeeStructureForm
from django.shortcuts import get_object_or_404, redirect
import base64
import datetime
import requests
from .mpesa import get_access_token
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponseForbidden, FileResponse
from django.utils import timezone
from datetime import timedelta
from .models import Payment
from django.core.paginator import Paginator

from .models import Student, Attendance
from datetime import date
from django.contrib.auth.models import User, Group
from django.db.models import Count, Q


from django.db.models.functions import TruncMonth
from .models import AttendanceAlert

from django.contrib.auth.decorators import permission_required

from accounts.models import TeacherProfile
from accounts.views import is_admin
from .utils import require_admin_approval
from .models import Payment, AdminActionLog
from django.contrib.auth import authenticate
from django.contrib.auth.decorators import user_passes_test

from .models import AcademicReport
from django.template.loader import get_template
from xhtml2pdf import pisa
from .models import AcademicReport, ReportSubject
from .forms import AcademicReportForm, ReportSubjectForm
from django.forms import inlineformset_factory
from django.contrib.admin.views.decorators import staff_member_required

from .utils import subject_performance
from django.contrib.auth.decorators import login_required, user_passes_test
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required

from .models import AcademicReport, Student
from students.models import Subject

from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import get_object_or_404, redirect, render
from django.forms import inlineformset_factory
from django.db.models import Prefetch

def is_staff_or_admin(user):
    return user.is_staff or user.is_superuser



# ====================== DASHBOARD ======================
@login_required
@user_passes_test(is_staff_or_admin)
def dashboard(request):
    
    cleanup_pending_payments()

    recent_payments = Payment.objects.filter(
        status="PAID"
    ).order_by('-date_paid')[:5]

    total_students = Student.objects.count()
    total_payments = Payment.objects.filter(status="PAID").count()
    amount_paid = Payment.objects.filter(status="PAID").aggregate(Sum("amount_paid"))["amount_paid__sum"] or 0
    total_amount_paid = amount_paid

    # ============================
    # EXPECTED FEES (BASED ON STUDENTS)
    # ============================
    total_expected_fees = 0
    fee_structures = FeeStructure.objects.all()

    for fee in fee_structures:
        num_students = Student.objects.filter(class_level=fee.class_name).count()
        total_expected_fees += fee.amount * num_students

    # ============================
    # PAYMENTS PER CLASS (BAR CHART)
    # ============================
    classes = []
    amount_per_class = []

    for fee in fee_structures:
        num_students = Student.objects.filter(class_level=fee.class_name).count()

        if num_students > 0:  # show only classes that have students
            classes.append(fee.class_name)

            total_in_class = Payment.objects.filter(
               student__class_level=fee.class_name,
               status="PAID"   # 🔥 HII NDIO FIX
            ).aggregate(total=Sum("amount_paid"))["total"] or 0


            amount_per_class.append(float(total_in_class))

    # ============================
    # STUDENTS WITH BALANCE
    # ============================
    students_with_balance = []

    students = Student.objects.all()
    for student in students:
        fee = FeeStructure.objects.filter(class_name=student.class_level).first()
        expected = fee.amount if fee else 0

        paid = Payment.objects.filter(student=student).aggregate(
            total=Sum("amount_paid")
        )["total"] or 0

        paid = Payment.objects.filter(student=student,
               status="PAID").aggregate(total=Sum("amount_paid"))["total"] or 0



        balance = expected - paid

        if balance > 0:
            students_with_balance.append({
                "name": student.full_name,
                "balance": float(balance)
            })

    # ============================
    # PIE CHART
    # ============================
    remaining_balance = total_expected_fees - total_amount_paid
    remaining_balance = max(remaining_balance, 0)

    context = {
        "total_students": total_students,
        "total_payments": total_payments,
        "total_amount_paid": float(total_amount_paid),

        # Replaced old wrong logic
        "total_fee_expected": float(total_expected_fees),

        "recent_payments": recent_payments,

        # Charts
        "classes": json.dumps(classes),
        "amount_per_class": json.dumps(amount_per_class),
        "students_with_balance": students_with_balance,
        "remaining_balance": float(remaining_balance),
    }

    return render(request, "dashboard.html", context)


# ====================== STUDENT PROFILE ======================
def student_profile(request, pk):
    student = get_object_or_404(Student, pk=pk)

    # ✅ Admin bypass
    if request.user.is_staff or request.user.is_superuser:
        verified = True
    else:
        verified = request.session.get(f"verified_student_{student.id}", False)
     

    payments = Payment.objects.filter(student=student)

    total_paid = Payment.objects.filter(
        student=student,
        status="PAID"
    ).aggregate(total=Sum("amount_paid"))["total"] or 0

    fee = FeeStructure.objects.filter(
        class_name=student.class_level
    ).first()

    total_fee = fee.amount if fee else 0
    balance = total_fee - total_paid

    
 

    return render(request, "student_profile.html", {
        "student": student,
        "payments": payments,
        "total_paid": total_paid,
        "total_fee": total_fee,
        "balance": balance,
        "verified": verified,  # 👈 muhimu
    })


def verify_admission(request, pk):
    student = get_object_or_404(Student, pk=pk)

    admission = request.POST.get("admission_number", "").strip()

    if admission == student.admission_number:
        # Save verification in session
        request.session[f"verified_student_{student.id}"] = True
        return JsonResponse({"success": True})

    return JsonResponse({
        "success": False,
        "message": "Invalid admission number. Please try again."
    })

# ====================== STUDENTS LIST ======================
def students_list(request):
    # --- BASE QUERY ---
    students = Student.objects.all()
    total_students = Student.objects.count()

       # --- GET FILTER VALUES (safe defaults & strip whitespace) ---
    search = request.GET.get("search", "").strip()
    class_level = request.GET.get("class", "").strip()   # we will pass this back as "class" in context
    start = request.GET.get("start", "").strip()
    end = request.GET.get("end", "").strip()
    # 🔐 ACCESS CONTROL LOGIC (MUHIMU SANA)
    # If NOT admin and NO filters → show NOTHING
    if not request.user.is_superuser:
        if not search and not class_level and not start and not end:
            students = Student.objects.none()
    # --- SEARCH (first_name OR last_name OR parent_name) ---
    if search:
        students = students.filter(
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(parent_name__icontains=search)
        )
    # --- FILTER BY CLASS (class_level field) ---
    if class_level:
        students = students.filter(class_level=class_level)

    # --- FILTER BY DATE (date_registered field) ---
    # handle: start+end, start only (>=), end only (<=)
    if start and end:
        students = students.filter(date_registered__range=[start, end])
    elif start:
        students = students.filter(date_registered__gte=start)
    elif end:
        students = students.filter(date_registered__lte=end)


            # ============================
    # ✅ PAGINATION HAPA 👇
    # ============================
    paginator = Paginator(students, 50)   # 50 students per page
    page_number = request.GET.get("page")
    students = paginator.get_page(page_number)

    # --- COMPUTE BALANCES (keep your original output structure) ---
    student_data = []
    

    for student in students:
        payments = Payment.objects.filter(student=student)
        total_paid = Payment.objects.filter(student=student,
                                            status="PAID").aggregate(total=Sum("amount_paid"))["total"] or 0
 

        print("STUDENT:", student.full_name, "CLASS:", student.class_level)

        fee = FeeStructure.objects.filter(class_name=student.class_level).first()

        print("FOUND FEE:", fee)

        total_fee = fee.amount if fee else 0
 
        balance = total_fee - total_paid

        student_data.append({
            "student": student,
            "balance": balance,
        })
    # --- PASS distinct class list for the dropdown (so your template loop works) ---
    classes = [c[0] for c in Student.CLASS_LEVELS]


    # --- CONTEXT (note: key name "class" preserved to match your template) ---
    context = {
        "student_data": student_data,
        "search": search,
        "class": class_level,
        "start": start,
        "end": end,
        "classes": classes,
        "total_students": total_students, 
    }

    return render(request, "students_list.html",context)
# ====================== DELETE STUDENT LIST ======================

from django.shortcuts import get_object_or_404, redirect, render
from .forms import StudentTeacherForm
from .models import Student

def manage_student_teachers(request, student_id):

    student = get_object_or_404(Student, id=student_id)

    if request.method == "POST":
        form = StudentTeacherForm(request.POST, instance=student)
        if form.is_valid():
            form.save()
            return redirect("students_list")

    else:
        form = StudentTeacherForm(instance=student)

    return render(request,
                  "manage_teachers_popup.html",
                  {"student": student, "form": form})


# ====================== DELETE STUDENT LIST ======================
@login_required
def delete_student(request, student_id):

    # Staff but NOT superuser → needs admin approval
    if request.user.is_staff and not request.user.is_superuser:
        if not request.session.get("admin_approved"):
            request.session["next_url"] = request.path
            messages.warning(request, "Admin approval required to delete.")
            return redirect("admin_approval")

    # Superuser OR approved staff
    student = get_object_or_404(Student, id=student_id)
    student.delete()

    # clear approval after action
    request.session.pop("admin_approved", None)

    messages.success(request, "Student deleted successfully.")
    return redirect("student_list")


# ====================== PAYMENTS LIST ======================
def payments_list(request):
    payments = Payment.objects.filter(
        status="PAID"
    ).select_related("student").order_by("-date_paid")

    return render(request, "payments_list.html", {"payments": payments})


# ====================== ADD PAYMENT ======================
def add_payment(request, pk):
    student = get_object_or_404(Student, pk=pk)

    if request.method == "POST":
        try:

            Payment.objects.create(
               student=student,
               amount_paid=request.POST.get("amount_paid"),
               method="Cash",
               status="PAID",
               notes=request.POST.get("notes", "")
            )

            messages.success(request, "Payment added successfully!")
        except:
            messages.error(request, "Error adding payment, try again!")

        return redirect("student_profile", pk=pk)

    return render(request, "add_payment.html", {"student": student})

# ====================== EDIT STUDENT ======================
def edit_student(request, pk):
    student = get_object_or_404(Student, pk=pk)

    if request.method == "POST":
        try:
            # Update basic fields
            student.first_name = request.POST.get("first_name")
            student.last_name = request.POST.get("last_name")
            student.class_level = request.POST.get("class_level")
            student.phone_number = request.POST.get("phone_number")
            student.parent_id_number = request.POST.get("parent_id_number")

            # Handle photo
            if "photo" in request.FILES:
                student.photo = request.FILES["photo"]

            # 🟡 GET FEE STRUCTURE YA CLASS MPYA
            fee_structure = FeeStructure.objects.filter(class_name=student.class_level).first()

            if fee_structure:
                student.balance = fee_structure.amount  # update student balance automatically

            student.save()

            messages.success(request, "Student updated successfully!")
            return JsonResponse({"success": True})

        except Exception as e:
            print(e)
            messages.error(request, "Failed to update student.")
            return JsonResponse({"error": "Error occurred while updating."})

    return JsonResponse({"error": "Invalid request"})


# ====================== STUDENT STATEMENT PDF ======================


def student_statement_pdf(request, student_id):

    # 1. GET STUDENT
    student = get_object_or_404(Student, id=student_id)

    # 2. GET PAYMENTS
    payments = Payment.objects.filter(student=student).order_by('date_paid')

    # 3. CALCULATIONS
    # Total fee kutoka FeeStructure (si lazima student awe na field total_fee)
    fee = FeeStructure.objects.filter(class_name=student.class_level).first()
    total_fee = fee.amount if fee else 0


    total_paid = payments.aggregate(total=Sum("amount_paid"))["total"] or 0
    balance = total_fee - total_paid

    # 4. FILE NAME & PATH
    file_name = f"{student.full_name.replace(' ', '_')}_Statement.pdf"
    file_path = os.path.join(settings.MEDIA_ROOT, file_name)

    # 5. GENERATE PDF (from utils/pdf_utils)
    generate_student_statement_pdf(
        file_path=file_path,
        student=student,
        payments =payments,
        total_paid=total_paid,
        total_fee=total_fee,
        balance=balance
    )

    # 6. RETURN PDF AS DOWNLOAD
    return FileResponse(
        open(file_path, 'rb'),
        as_attachment=True,
        filename=file_name
    )

# ============================================

def edit_payment(request, payment_id):
    payment = get_object_or_404(Payment, id=payment_id)

    if request.method == "POST":
        payment.amount_paid = request.POST.get("amount_paid")
        payment.method = request.POST.get("method")
        payment.save()
        return redirect("payments_list")

    return render(request, "edit_payment.html", {"payment": payment})

# ====================== DELETE ======================

@login_required
def delete_payment(request, payment_id):

    # 1️⃣ Only staff or superuser can even attempt
    if not request.user.is_staff:
        messages.error(request, "You are NOT allowed to delete payments.")
        return redirect("payments_list")

    # 2️⃣ Staff (not superuser) must get admin approval
    if not require_admin_approval(request):
        return redirect("admin_approval")

    # 3️⃣ Delete payment
    payment = get_object_or_404(Payment, id=payment_id)
    payment.delete()

    # 4️⃣ Log the action (VERY IMPORTANT)
    AdminActionLog.objects.create(
        user=request.user,
        action=f"Deleted payment ID {payment_id}"
    )

    # 5️⃣ Clear approval after action
    request.session.pop("admin_approved", None)

    messages.success(request, "Payment deleted successfully.")
    return redirect("payments_list")

# ====================== API ======================

def api_payment(request, payment_id):
    payment = get_object_or_404(Payment, id=payment_id)
    return JsonResponse({
        "id": payment.id,
        "amount_paid": payment.amount_paid,
        "method": payment.method
    })


def api_payment_details(request, payment_id):
    try:
        payment = Payment.objects.get(id=payment_id)
        data = {
            "id": payment.id,
            "amount_paid": float(payment.amount_paid),   # 👈 BADILISHA HAPA
            "method": payment.method,                   # 👈 ADD THIS
        }
        return JsonResponse(data)
    except Payment.DoesNotExist:
        return JsonResponse({"error": "Payment not found"}, status=404)

  

  # students/views.py



def download_receipt(request, payment_id):
    payment = get_object_or_404(Payment, id=payment_id)

    # 🚫 BLOCK DOWNLOAD IF NOT PAID
    if payment.status != "PAID":
        return HttpResponseForbidden(
            "Receipt can only be downloaded for PAID payments."
        )

    student = payment.student

    file_name = f"Receipt_{payment_id}.pdf"
    file_path = os.path.join(settings.MEDIA_ROOT, file_name)

    # ✅ Generate the receipt PDF
    generate_receipt_pdf(file_path, student, payment)

    # ✅ Return PDF to browser
    return FileResponse(
        open(file_path, 'rb'),
        as_attachment=True,
        filename=file_name
    )



def admin_required(view_func):
    decorated_view_func = user_passes_test(
        lambda user: user.is_staff,    # condition = lazima awe admin
        login_url='dashboard'          # akishindwa → mpeleke dashboard
    )(view_func)
    return decorated_view_func



@login_required
def student_add(request):
    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES)

        if form.is_valid():
            student = form.save(commit=False)   # ⚠️ bado haijasave DB
            student.save()                      # sasa main student anaingia

            # ✅ SAVE MANY-TO-MANY TEACHERS
            form.save_m2m()

            return redirect('student_list')
    else:
        form = StudentForm()

    return render(request, 'student_add.html', {'form': form})



def add_feestructure_popup(request):
    if request.method == 'POST':
        form = FeeStructureForm(request.POST)
        if form.is_valid():
            obj = form.save()
            return render(request, 'popup_close.html', {'object': obj})
    else:
        form = FeeStructureForm()

    return render(request, 'add_feestructure_popup.html', {'form': form})


@login_required
def feestructure_list(request):
    items = FeeStructure.objects.all().order_by('-created_at')
    return render(request, 'feestructure_list.html', {'items': items})



def public_feestructure(request):
    items = FeeStructure.objects.all().order_by('class_name')
    return render(request, 'public_feestructure.html', {'items': items})


def feestructure_edit(request, pk):
    feestructure = get_object_or_404(FeeStructure, pk=pk)

    if request.method == 'POST':
        form = FeeStructureForm(request.POST, instance=feestructure)
        if form.is_valid():
            form.save()
            return redirect('feestructure_list')
    else:
        form = FeeStructureForm(instance=feestructure)

    return render(request, 'feestructure_edit.html', {'form': form})


def feestructure_delete(request, pk):
    feestructure = get_object_or_404(FeeStructure, pk=pk)

    if request.method == 'POST':
        feestructure.delete()
        return redirect('feestructure_list')

    return render(request, 'feestructure_confirm_delete.html', {'fee': feestructure})


def mpesa_payment(request, student_id):

    if request.method != "POST":
        return redirect("student_profile", pk=student_id)

    student = get_object_or_404(Student, id=student_id)

    phone = request.POST.get("phone", "").strip()
    amount = int(request.POST.get("amount", 0))

    if phone.startswith("+"):
        phone = phone[1:]
    if phone.startswith("0"):
        phone = "254" + phone[1:]

    access_token = get_access_token()

    timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    password_str = settings.MPESA_SHORTCODE + settings.MPESA_PASSKEY + timestamp
    password = base64.b64encode(password_str.encode()).decode()

    account_reference = f"STUDENT-{student.id}-{timestamp}"

    payload = {
        "Password": "MTc0Mzc5YmZiMjc5ZjlhYTliZGJjZjE1OGU5N2RkNzFhNDY3Y2QyZTBjODkzMDU5YjEwZjc4ZTZiNzJhZGExZWQyYzkxOTIwMjUxMjE2MTQ1MDQ4",
        "BusinessShortCode": "4189211",
        "Timestamp": "20251216145048",
        "Amount": "1",
        "PartyA": "254708374149",
        "PartyB": "4189211",
        "TransactionType": "CustomerPayBillOnline",
        "PhoneNumber": "254708374149",
        "CallBackURL":  "https://paz-eustatic-stridently.ngrok-free.dev/mpesa/callback/",
        "AccountReference": account_reference,
        "TransactionDesc": "StudentsCo Fees Payment"
    }

    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json"
    }

    stk_url = (
        "https://api.safaricom.co.ke/mpesa/stkpush/v1/processrequest"
        if settings.MPESA_ENV == "live"
        else "https://sandbox.safaricom.co.ke/mpesa/stkpush/v1/processrequest"
    )

    response = requests.post(stk_url, json=payload, headers=headers)

    print("STK STATUS:", response.status_code)
    print("STK RESPONSE:", response.text)

    data = response.json()


    if data.get("ResponseCode") != "0":
         messages.error(request, "STK Push failed")
         return redirect("student_profile", pk=student.id)

    # ✅ STK SUCCESS → HAPA NDIPO CHECKOUT INACHUKULIWA
    checkout_request_id = data.get("CheckoutRequestID")

    Payment.objects.create(
    student=student,
    amount_paid=amount,
    method="Mpesa",
    status="PENDING",
    checkout_request_id=checkout_request_id,
    notes=f"Ref:{account_reference}|Phone:{phone}"
)

    messages.success(request, "Check your phone and enter M-Pesa PIN")
    return redirect("student_profile", pk=student.id)

@csrf_exempt
def mpesa_callback(request):

    # Allow Safaricom GET ping
    if request.method != "POST":
        return HttpResponse("M-PESA Callback Endpoint", status=200)

    data = json.loads(request.body.decode("utf-8"))
    print("MPESA CALLBACK DATA:", data)

    stk_callback = data.get("Body", {}).get("stkCallback", {})
    result_code = stk_callback.get("ResultCode")
    result_desc = stk_callback.get("ResultDesc")
    checkout_request_id = stk_callback.get("CheckoutRequestID")

    payment = Payment.objects.filter(
        checkout_request_id=checkout_request_id
    ).first()

    if not payment:
        return JsonResponse({"ResultCode": 0, "ResultDesc": "No pending payment"})

    if result_code == 0:
        metadata = stk_callback.get("CallbackMetadata", {}).get("Item", [])

        receipt = ""
        phone = ""

        for item in metadata:
            if item["Name"] == "MpesaReceiptNumber":
                receipt = item["Value"]
            elif item["Name"] == "PhoneNumber":
                phone = item["Value"]

        # ✅ SUCCESS
        payment.status = "PAID"
        payment.notes = f"CONFIRMED | Receipt: {receipt} | Phone: {phone}"
        payment.save()

    else:
       
        payment.status = "FAILED"
        payment.notes = f"FAILED | {result_desc}"
        payment.save()

    return JsonResponse({"ResultCode": 0, "ResultDesc": "Accepted"})


def cleanup_pending_payments():
    expiry = timezone.now() - timedelta(minutes=2)
    Payment.objects.filter(
        status="PENDING",
        created_at__lt=expiry
    ).delete()   # AU .update(status="FAILED")


@login_required
def brand_intro(request):
    return render(request, 'brand_intro.html')



@login_required
def teacher_dashboard(request):

    # 🔐 Teacher only
    if not hasattr(request.user, "teacherprofile"):
        return render(request, "403.html")

    user = request.user


    today = date.today()

    # 👨‍🎓 TOTAL STUDENTS
  
    total_students = Student.objects.filter(
    Q(teacher=user) |
    Q(extra_teachers=user)
).distinct().count()

    # 📝 MARKED TODAY
    marked_today = Attendance.objects.filter(
        teacher=user,
        date=today
    ).count()

    # 🔑 Attendance status logic
    if marked_today == 0:
        attendance_status = "not_started"
    elif marked_today < total_students:
        attendance_status = "incomplete"
    else:
        attendance_status = "completed"

    # 📊 Attendance percentage
    if total_students > 0:
        attendance_percentage = int((marked_today / total_students) * 100)
    else:
        attendance_percentage = 0

    # ==============================
    # 📊 PERFORMANCE LOGIC (ADD HERE)
    # ==============================

    attendance_qs = Attendance.objects.filter(teacher=user)

    # ⚠ At-risk students (4+ absents)
    at_risk_students = (
        attendance_qs
        .filter(status="Absent")
        .values("student")
        .annotate(absent_count=Count("id"))
        .filter(absent_count__gte=4)
    )

    at_risk_count = at_risk_students.count()

    # 🏆 Top performer (least absents)
    top_performer = (
        attendance_qs
        .values(
            "student__first_name",
            "student__last_name"
        )
        .annotate(
            absent_count=Count(
                "id",
                filter=Q(status="Absent")
            )
        )
        .order_by("absent_count")
        .first()
    )

    new_alerts_count = AttendanceAlert.objects.filter(
    teacher=request.user,
    status="new").count()



    # ==============================
    # CONTEXT
    # ==============================

    context = {
        "total_students": total_students,
        "marked_today": marked_today,
        "attendance_status": attendance_status,
        "attendance_percentage": attendance_percentage,
        "at_risk_count": at_risk_count,
        "top_performer": top_performer,
        "new_alerts_count": new_alerts_count,
    }

    return render(request, "teacher_dashboard.html", context)


@login_required
def at_risk_students(request):

    # 🔐 Teacher only
    if not hasattr(request.user, "teacherprofile"):
       return render(request, "403.html")

    teacher = request.user

    # 1️⃣ Get students with 4+ absents
    risky_students = (
        Attendance.objects
        .filter(teacher=teacher, status="Absent")
        .values("student")
        .annotate(absent_count=Count("id"))
        .filter(absent_count__gte=4)
    )

    # 2️⃣ Ensure alert exists for each at-risk student
    for s in risky_students:
        AttendanceAlert.objects.get_or_create(
            teacher=teacher,
            student_id=s["student"],
            defaults={"status": "new"}
        )

    # 3️⃣ Fetch alerts (THIS replaces your old query)
    students = (
        AttendanceAlert.objects
        .filter(teacher=teacher)
        .select_related("student")
        .order_by("-created_at")
    )

    context = {
        "students": students
    }

    return render(request, "at_risk_students.html", context)


@login_required
def mark_alert_reviewed(request, alert_id):
    alert = AttendanceAlert.objects.get(
        id=alert_id,
        teacher=request.user
    )
    alert.status = "reviewed"
    alert.save()

    return redirect("at_risk_students")


   
from django.db.models import Q

@login_required
def teacher_students(request):

    if not hasattr(request.user, "teacherprofile"):
       return render(request, "403.html")



    students = Student.objects.filter(
    Q(teacher=request.user) |
    Q(extra_teachers=request.user)
).distinct()



    context = {
        "students": students,
        "total_students": students.count()
    }

    return render(request, "my_students.html", context)




@login_required
def mark_attendance(request):

    # 🔐 Teacher only
    if not hasattr(request.user, "teacherprofile"):
       return render(request, "403.html")


    teacher = request.user
    students = Student.objects.filter(
    Q(teacher=teacher) |
    Q(extra_teachers=teacher)
).distinct()

    today = timezone.localdate()

    total_students = students.count()

    marked_today = Attendance.objects.filter(
        teacher=teacher,
        date=today
    ).count()

    # 🔒 LOCK LOGIC (HAPA NDIPO PANAWEKWA)
    attendance_locked = marked_today == total_students and total_students > 0

    # 🚫 BLOCK POST if locked
    if request.method == "POST" and attendance_locked:
        messages.warning(
            request,
            "Attendance for today is already completed and locked."
        )
        return redirect("teacher_attendance")

    # ✅ SAVE ATTENDANCE
    if request.method == "POST":
        for student in students:
            status = request.POST.get(f"status_{student.id}")

            if status:
                Attendance.objects.update_or_create(
                    student=student,
                    date=today,
                    defaults={
                        "teacher": teacher,
                        "status": status,
                        "is_locked": True
                    }
                )

        messages.success(request, "Attendance saved successfully.")
        return redirect("teacher_attendance")

    # 🧠 Load existing attendance
    attendance_map = {
        a.student_id: a.status
        for a in Attendance.objects.filter(
            teacher=teacher,
            date=today
        )
    }

    context = {
        "students": students,
        "today": today,
        "attendance_map": attendance_map,
        "attendance_locked": attendance_locked,
    }

    return render(request, "attendance.html", context)


@login_required
def attendance_history(request):

    # Teacher only
    if not hasattr(request.user, "teacherprofile"):
       return render(request, "403.html")


    teacher = request.user

    # Selected date
    selected_date = request.GET.get("date")
    if selected_date:
        selected_date = date.fromisoformat(selected_date)
    else:
        selected_date = timezone.localdate()

    # Attendance records for selected date
    records = (
        Attendance.objects
        .filter(teacher=teacher, date=selected_date)
        .select_related("student")
    )

    # 🚨 Students with 4+ absents
    absent_alerts = (
        Attendance.objects
        .filter(teacher=teacher, status="Absent")
        .values(
            "student_id",
            "student__first_name",
            "student__last_name"
        )
        .annotate(absent_count=Count("id"))
        .filter(absent_count__gte=4)
    )

    context = {
        "records": records,
        "selected_date": selected_date,
        "absent_alerts": absent_alerts,
    }

    return render(request, "attendance_history.html", context)


@login_required
def monthly_attendance_summary(request):

    # 🔐 Teacher only
    if not hasattr(request.user, "teacherprofile"):
       return render(request, "403.html")


    teacher = request.user

    # 📅 Selected month
    selected_month = request.GET.get("month")

    if selected_month:
        year, month = map(int, selected_month.split("-"))
        start_date = date(year, month, 1)
    else:
        today = timezone.now().date()
        start_date = date(today.year, today.month, 1)

    # 📊 Attendance summary per student
    summary = (
        Attendance.objects
        .filter(
            teacher=teacher,
            date__year=start_date.year,
            date__month=start_date.month
        )
        .values(
            "student_id",
            "student__first_name",
            "student__last_name"
        )
        .annotate(
            total_days=Count("id"),
            absent_days=Count("id", filter=Q(status="Absent"))
        )
        .order_by("student__first_name")
    )

    # 📈 Add attendance percentage
    for s in summary:
        if s["total_days"] > 0:
            s["attendance_percentage"] = int(
                ((s["total_days"] - s["absent_days"]) / s["total_days"]) * 100
            )
        else:
            s["attendance_percentage"] = 0

    context = {
        "summary": summary,
        "selected_month": start_date.strftime("%Y-%m"),
    }

    return render(
        request,
        "monthly_attendance_summary.html",
        context
    )



@login_required
def admin_attendance_overview(request):
    if not request.user.is_staff:
        return render(request, "403.html")

    records = (
        Attendance.objects



        .select_related(
                        "student",
                        "student__feestructure",
                        "student__teacher",
                        "teacher"
                                 )


        .order_by("-date")
    )

    return render(
        request,
        "attendance_overview.html",
        {"records": records}
    )


@login_required
def admin_locked_attendance(request):
    if not request.user.is_staff:
        return render(request, "403.html")

    records = Attendance.objects.filter(is_locked=True)

    return render(
        request,
        "locked_attendance.html",
        {"records": records}
    )

@login_required
def unlock_attendance(request, attendance_id):
    if not request.user.is_staff:
        return render(request, "403.html")

    attendance = get_object_or_404(Attendance, id=attendance_id)
    attendance.is_locked = False
    attendance.save()

    messages.success(request, "Attendance unlocked successfully.")
    return redirect("locked_attendance")



@login_required
def attendance_alerts(request):
    if not request.user.is_staff:
        return render(request, "403.html")

    alerts = (
        AttendanceAlert.objects
        .select_related("student", "teacher")
        .order_by("-created_at")
    )

    return render(
        request,
        "attendance_alerts.html",
        {"alerts": alerts}
    )



def home_dashboard(request):
    return render(request, "home.html")



def search_student(request):
    q = request.GET.get('q', '').strip()

    if len(q) < 2:
        return JsonResponse({'found': False})

    parts = q.split()

    if len(parts) == 1:
        student = Student.objects.filter(
            Q(first_name__icontains=parts[0]) |
            Q(last_name__icontains=parts[0])
        ).first()
    else:
        student = Student.objects.filter(
            Q(first_name__icontains=parts[0]) &
            Q(last_name__icontains=parts[-1])
        ).first()

    if student:
        return JsonResponse({
            'found': True,
            'name': f"{student.first_name} {student.last_name}",
            'message': 'Student is registered. Scroll down to view full details.'
        })

    return JsonResponse({'found': False})




def admin_approval(request):

    logs = AdminActionLog.objects.order_by("-timestamp")[:10]

    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        admin = authenticate(request, username=username, password=password)

        if admin and admin.is_superuser:
            request.session["admin_approved"] = True

            # log approval
            AdminActionLog.objects.create(
                user=admin,
                action=f"Approved action for {request.user.username}"
            )

            messages.success(request, "Admin approval granted.")
            return redirect(request.session.get("next_url", "dashboard"))

        messages.error(request, "Invalid admin credentials.")

    return render(request, "admin_approval.html", {
        "logs": logs
    })


@user_passes_test(lambda u: u.is_superuser)
def delete_admin_log(request, log_id):

    log = get_object_or_404(AdminActionLog, id=log_id)
    log.delete()

    # 🔥 VERY IMPORTANT
    request.session.pop("admin_approved", None)

    messages.success(request, "Log deleted successfully.")
    return redirect("admin_approval")




@login_required
def student_reports(request):
    term = request.GET.get("term")        # T1, T2, T3
    exam_type = request.GET.get("exam")   # MID, END

    reports = AcademicReport.objects.filter(status="PUBLISHED")

    if term:
        reports = reports.filter(term=term)

    if exam_type:
        reports = reports.filter(exam_type=exam_type)

    reports = (
    reports
    .select_related("student")
    .prefetch_related("subjects")
    .order_by(
        "student__class_level",
        "-total_score",
        "-mean_marks"
    )
)



    classes = {}

    for report in reports:
       class_name = report.student.class_level
       classes.setdefault(class_name, []).append(report)

             # calculate position per class
    for class_name, class_reports in classes.items():
       sorted_reports = sorted(
        class_reports,
        key=lambda r: (r.total_score, r.mean_marks),
        reverse=True
    )

    for index, report in enumerate(sorted_reports, start=1):
        report.position = index
        


    return render(
        request,
        "student_reports.html",
        {
            "classes": classes,
            "selected_term": term,
            "selected_exam": exam_type,
        }
    )



@staff_member_required
def admin_reports(request):
    term = request.GET.get("term")
    exam_type = request.GET.get("exam")

    reports = AcademicReport.objects.all()
    

    if term:
        reports = reports.filter(term=term)

    if exam_type:
        reports = reports.filter(exam_type=exam_type)

    return render(request, "reports_list.html", {
        "reports": reports
    })




@staff_member_required
def publish_report(request, pk):
    report = get_object_or_404(AcademicReport, pk=pk)

    if report.status == "PUBLISHED":
        report.status = "DRAFT"
    else:
        report.status = "PUBLISHED"

    report.save()
    return redirect("admin_reports")


# ================================
# helpers (safe kabisa)
# ================================
# ================================
def get_subject_grade(marks):
    marks = float(marks or 0) 
    if marks >= 80: return "A"
    if marks >= 75: return "A-"
    if marks >= 70: return "B+"
    if marks >= 65: return "B"
    if marks >= 60: return "B-"
    if marks >= 55: return "C+"
    if marks >= 50: return "C"
    if marks >= 45: return "C-"
    if marks >= 40: return "D+"
    if marks >= 35: return "D"
    if marks >= 30: return "D-"
    return "E"


def get_report_grade(avg):
    avg = float(avg or 0)   # 🔒 SAFE FIX
    if avg >= 80: return "A"
    if avg >= 75: return "A-"
    if avg >= 70: return "B+"
    if avg >= 65: return "B"
    if avg >= 60: return "B-"
    if avg >= 55: return "C+"
    if avg >= 50: return "C"
    if avg >= 45: return "C-"
    if avg >= 40: return "D+"
    if avg >= 35: return "D"
    if avg >= 30: return "D-"
    return "E"
@staff_member_required
def add_report(request, student_id):
    student = get_object_or_404(Student, pk=student_id)

    ReportSubjectFormSet = inlineformset_factory(
        AcademicReport,
        ReportSubject,
        form=ReportSubjectForm,
        extra=5,
        can_delete=False
    )

    if request.method == "POST":
        report_form = AcademicReportForm(request.POST)
        formset = ReportSubjectFormSet(
            request.POST,
            form_kwargs={"student": student}
        )

        if report_form.is_valid() and formset.is_valid():
            # 1️⃣ save report
            report = report_form.save(commit=False)
            report.student = student
            report.total_score = 0   # 🔥 ADD THIS
            report.grade = ""        # optional but safe
            report.save()

            # 2️⃣ save subjects
            formset.instance = report
            formset.save()

            # 3️⃣ calculations (EXCEL STYLE 🔥)
            subjects = report.subjects.all()

            total = 0
            for s in subjects:
                s.grade = get_subject_grade(s.marks)
                total += s.marks
                s.save()

            avg = total / subjects.count() if subjects.exists() else 0

            report.total_score = total
            report.mean_marks = avg     
            report.grade = get_report_grade(avg)
            report.save()

            return redirect("admin_reports")

    else:
        report_form = AcademicReportForm()
        formset = ReportSubjectFormSet(
            form_kwargs={"student": student}
        )

    return render(request, "add_report.html", {
        "report_form": report_form,
        "formset": formset,
        "student": student
    })


@staff_member_required
def edit_report(request, pk):
    report = get_object_or_404(AcademicReport, pk=pk)

    ReportSubjectFormSet = inlineformset_factory(
        AcademicReport,
        ReportSubject,
        form=ReportSubjectForm,
        extra=0,
        can_delete=False
    )

    if request.method == "POST":
        report_form = AcademicReportForm(request.POST, instance=report)
        formset = ReportSubjectFormSet(
            request.POST,
            instance=report,
            form_kwargs={"student": report.student}
        )

        if report_form.is_valid() and formset.is_valid():
            report_form.save()
            formset.save()
            return redirect("admin_reports")
    else:
        report_form = AcademicReportForm(instance=report)
        formset = ReportSubjectFormSet(
            instance=report,
            form_kwargs={"student": report.student}
        )

    return render(request, "edit_report.html", {
        "report_form": report_form,
        "formset": formset,
        "report": report
    })




@login_required
def download_report_pdf(request, pk):
    report = get_object_or_404(
        AcademicReport,
        pk=pk,
        status="PUBLISHED"

    )
    previous = (
    AcademicReport.objects
    .filter(
        student=report.student,
        status="PUBLISHED"
    )
    .exclude(pk=report.pk)
    .order_by("-created_at")
    .first()
)


    performance = subject_performance(report)

    template = get_template("report_pdf.html")
    html = template.render({
        "report": report,
        "performance": performance,
        "previous": previous,
    })

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="report_{report.student.admission_number}.pdf"'
    )

    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse("PDF generation error")

    return response



@staff_member_required
def delete_report(request, pk):
    report = get_object_or_404(AcademicReport, pk=pk)

    report.delete()
    messages.success(request, "Academic report deleted successfully.")

    return redirect("admin_reports")


@login_required
def export_class_broadsheet(request):
    wb = Workbook()
    wb.remove(wb.active)

    # ===== Styles =====
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    center = Alignment(horizontal="center", vertical="center")

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ===== Subjects (fixed order – school standard) =====
    all_subjects = [
        "English", "Kiswahili", "Mathematics", "Biology", "Physics",
        "Chemistry", "Geography", "C.R.E", "Agriculture",
        "Computer Studies", "Business Studies", "History"
    ]

    # ===== Classes =====
    classes = Student.objects.values_list("class_level", flat=True).distinct()

    for class_name in classes:
        ws = wb.create_sheet(title=str(class_name))

        # ===== Title =====
        total_columns = 2 + len(all_subjects) + 4  # ADM, NAME + subjects + TOTAL, AVG, GRADE, POS
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
        ws.cell(row=1, column=1).value = f"{class_name} – Academic Broadsheet"
        ws.cell(row=1, column=1).font = title_font
        ws.cell(row=1, column=1).alignment = center

        # ===== Headers =====
        headers = ["ADM", "NAME"] + all_subjects + ["TOTAL", "AVG", "GRADE", "POS"]
        ws.append(headers)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=2, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
            ws.column_dimensions[get_column_letter(col)].width = 15

        # ===== Reports (ORDER MATTERS) =====
        reports = (
            AcademicReport.objects
            .select_related("student")
            .prefetch_related("subjects")
            .filter(
                student__class_level=class_name,
                status="PUBLISHED"
            )
            .order_by("-total_score", "-mean_marks")
        )

        # ===== Rows + Position =====
        row = 3
        position = 1

        for report in reports:
            student = report.student

            subjects = {
                (s.subject.name if s.subject else ""): s.marks
                for s in report.subjects.all()
            }

            row_data = [
                student.admission_number,
                f"{student.first_name} {student.last_name}",
            ]

            for sub in all_subjects:
                row_data.append(subjects.get(sub, ""))

            row_data += [
                report.total_score,
                round(report.mean_marks, 2),
                report.grade,
                position,
            ]

            ws.append(row_data)

            for col in range(1, len(row_data) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                cell.alignment = center

            position += 1
            row += 1

    # ===== Response =====
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Academic_Broadsheet.xlsx"'
    wb.save(response)
    return response
